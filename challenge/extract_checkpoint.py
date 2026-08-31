"""Extract the combined audit checkpoint workbook into committed CSVs.

The workbook arrived 2026-08-30, after Freeze 1, blind coding, and adjudication
were complete. It is an INDEPENDENT RECONSTRUCTION of the same job search,
produced separately from this repository and holding 353 records against this
repository's 298. It is therefore treated as an adversarial challenger rather
than as a coder: it saw a different source set, so it is not a third blind
rating and no kappa is computed against it.

Two things in it are primary evidence rather than reconstruction, and only
those enter the corpus:

  * `LinkedIn Applications`, 107 rows from the real LinkedIn data download
    (`Job Applications_5.csv` and `_6.csv`), carrying job IDs, job URLs, and
    EXACT application dates. This is the artifact stop condition 3 has been
    waiting for. It supersedes the 99-row paged scrape, whose stamps were
    relative.
  * `Jobright Applications`, 40 rows, already in the corpus at Freeze 2.

Everything else is written under `challenge/` and is never read by the census
pipeline. It exists so the reconciliation in `challenge/CHALLENGE.md` can be
recomputed rather than asserted.

REDACTION. The workbook names real recruiters and other third parties, far more
identifiably than the existing corpus does, which stores at most a first name in
`counterparty_name`. Committed output here replaces every personal name with a
stable one-way pointer `per_<sha256[:12]>`, following the convention in
`scripts/redact_corpus.py`. Company names stay, because the study is about
companies. No reverse map is committed.

    python3 challenge/extract_checkpoint.py path/to/workbook.xlsx

The workbook itself is deliberately NOT committed. It contains unredacted
personal names and this repository is public. Its sha256 is recorded below so
provenance is checkable by anyone holding the file.
"""

from __future__ import annotations

import csv
import hashlib
import os
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "challenge"
ARTIFACTS = ROOT / "artifacts" / "platform"

# Successive versions of the challenger, newest first. The workbook is revised
# between passes, so the extractor accepts any known version and names which one
# it read rather than failing on a hash it has not seen.
KNOWN_SOURCES = {
    "d32c869ad113a32cdde646ab9fb6a76336d22284935b66049e238cb40427b589": "combined_job_search_audit_checkpoint_refined.xlsx (343 records, register sheet)",
    "5a5c012f3a438ef388ba5afb235d635ecdffb860c9a7979f647a581db402c0a9": "combined_job_search_audit_checkpoint_updated.xlsx (353 records)",
}
SOURCE_SHA256 = "d32c869ad113a32cdde646ab9fb6a76336d22284935b66049e238cb40427b589"
SOURCE_NAME = "combined_job_search_audit_checkpoint_refined.xlsx"
# The workbook's own stated freshness boundary. Activity after this date is not
# in it, which bounds every coverage claim the challenge makes.
EXPORT_FRESHNESS = "2026-08-23"


# Phrases the harvests must never treat as a person.
KNOWN_NON_PERSONS = {
    "Clay Café", "Clay Cafe", "GTM Café", "GTM Cafe", "Work at a Startup",
    "Easy Apply", "Saved Jobs", "Job Applications", "United States",
    "New York", "San Francisco", "CRO idea",
}

# Single capitalized words that are never a given name in this workbook. Kept
# separate from KNOWN_NON_PERSONS because the single-token rule is the loose one
# and needs its own stop list.
NON_NAME_WORDS = {
    "unknown", "none", "no", "n/a", "na", "tbd", "various", "multiple", "yes",
    "unresolved", "pending", "unnamed", "unidentified", "recruiter", "founder",
    "ceo", "cto", "cro", "vp", "interviewer", "interview", "interviews",
    "contact", "contacts", "round", "rounds", "self", "team", "panel", "hiring",
    "manager", "director", "head", "lead", "unclear", "other",
}

# Given names that appear ONLY in prose, where no split rule isolates them and a
# single capitalized word cannot be told from a sentence-initial common word.
# This list is the output of the review guard below, classified by hand, and the
# guard fails the extraction when it finds a candidate that is on neither list.
# A name here is redacted wherever it appears, in any column.
PROSE_GIVEN_NAMES = {
    "Patrick", "Kellen", "Eoin",
}

# Capitalized words seen in this workbook's prose that are not people. Reviewed
# once, recorded here so the guard stays quiet about them and loud about
# anything new. Three kinds: sentence words, company names that must survive
# redaction because the study is about companies, and the tools the workbook
# names as evidence sources.
PROSE_NON_NAMES = {
    # Sentence words.
    "Calendar", "Community", "Company", "Completed", "Contact", "Do", "First",
    "Interview", "Interviewer", "Job", "Not", "One", "Paid", "Real", "Take",
    "Track", "Two", "No", "Second", "Prior", "Existing", "Exact", "Title",
    "Preserve", "Three", "Both", "Only", "Its", "The", "This", "That", "There",
    "These", "Where", "When", "While", "After", "Before", "Because", "If",
    "Alignment", "Cybersecurity", "Direct", "Eight", "Engineer", "Facilitated",
    "Fifteen", "Good", "Met", "Receipts", "Work",
    # Companies. A word here is a company this study reports on, and hashing it
    # would destroy the finding rather than protect anyone.
    "Apollo", "Beautiful", "Blackthorn", "Cargo", "Kivira", "Morph", "Numeric",
    "Pin",
    # Tools and platforms the workbook cites as evidence sources.
    "Ashby", "Gem", "Gmail", "Jobright", "Wellfound", "Zoom",
}

# Tokens that mark an organization or a job title rather than a person. Used
# ONLY by the prose harvest, which is a heuristic net over free text. The
# explicit person columns are not filtered this way: those hold people by
# definition, and blocking a token there would leave a real name unredacted,
# which is the error this module exists to prevent.
ORG_TOKENS = {
    "school", "schools", "college", "university", "district", "county", "city",
    "public", "global", "business", "development", "engineering", "technology",
    "technologies", "group", "security", "labs", "systems", "solutions",
    "services", "health", "capital", "ventures", "partners", "strategies",
    "data", "software", "media", "digital", "consulting", "recruiting",
    "talent", "startup", "cafe", "café", "team", "studio", "collective",
}


def load_csv_column(relative: str, columns: tuple[str, ...]) -> set[str]:
    """Values from named columns of a committed CSV, for the organization guard.

    Returns empty when the file is absent rather than failing: the extractor must
    still run on a fresh checkout that has not built the census yet. Losing this
    source over-redacts, which is the safe direction.
    """
    path = ROOT / relative
    if not path.is_file():
        return set()
    values: set[str] = set()
    with path.open(newline="", encoding="utf-8") as handle:
        for row in csv.DictReader(handle):
            for column in columns:
                value = (row.get(column) or "").strip()
                if value:
                    values.add(value)
    return values


def person_pointer(name: str) -> str:
    return "per_" + hashlib.sha256(name.strip().lower().encode()).hexdigest()[:12]


def load_workbook(path: Path):
    try:
        import openpyxl
    except ImportError:
        raise SystemExit(
            "openpyxl is required to re-extract the workbook.\n"
            "The extracted CSVs under challenge/ are committed, so the rest of the\n"
            "pipeline runs without it. Install it only to re-run this step."
        )
    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def sheet_rows(wb, name: str) -> tuple[list[str], list[dict]]:
    """Find the header row, then read data beneath it.

    Most sheets put a banner on row 1, the header on row 2, data from row 3. The
    Interview and Opportunity Register added in the refined workbook carries a
    banner, a summary strip, and a scope note before its header, so a hardcoded
    row index reads its header as data.

    A banner is recognized by its shape rather than its position: the workbook
    writes a merged title, which openpyxl returns as the SAME string repeated
    across every cell in the range. A header row has distinct values.
    """
    rows = [r for r in wb[name].iter_rows(values_only=True)]
    header_index = 1
    for i, row in enumerate(rows[:12]):
        filled = [str(c).strip() for c in row if c not in (None, "")]
        if len(set(filled)) >= 3 and i + 1 < len(rows) and any(c not in (None, "") for c in rows[i + 1]):
            header_index = i
            break
    header = [str(c).strip() if c is not None else "" for c in rows[header_index]]
    data = [
        dict(zip(header, r))
        for r in rows[header_index + 1 :]
        if any(c not in (None, "") for c in r)
    ]
    if not [h for h in header if h]:
        raise SystemExit(
            f"sheet {name!r}: no header row found in the first 12 rows. "
            "Header detection failed rather than silently writing empty columns."
        )
    return header, data


def clean(value) -> str:
    """Normalize a cell to text. Datetimes become ISO dates, not timestamps."""
    if value is None:
        return ""
    text = str(value).strip()
    if re.fullmatch(r"\d{4}-\d{2}-\d{2} 00:00:00", text):
        return text[:10]
    return text


def looks_like_person(part: str, organizations: set[str]) -> bool:
    """One shared shape test, so the two harvests cannot drift apart.

    Every token must start uppercase, which is what stops "CRO idea" being
    rostered. Whitespace is normalized by the caller, because a roster key with
    a doubled space can never match the text it came from and the name then
    ships unredacted.
    """
    tokens = part.split()
    if not 2 <= len(tokens) <= 3:
        return False
    if part in KNOWN_NON_PERSONS or part.isupper():
        return False
    if not all(token[:1].isupper() for token in tokens):
        return False
    # Substring rather than equality: "Atlanta Public Schools" must be protected
    # by an organization recorded as "Atlanta Public Schools, GA".
    #
    # The reverse direction is length-gated. The workbook holds a company called
    # "Vi", and an ungated `org in part` test made that two-letter string block
    # every name containing it: "Teresa Vitale" and "Vikas CV" both shipped in
    # the clear because of it. Four characters is the same floor `company_key`
    # uses before it stops stripping.
    if any(part in org for org in organizations):
        return False
    if any(org in part for org in organizations if len(org) >= 4):
        return False
    # Applied to BOTH harvests. The explicit columns were meant to hold people
    # only, but the workbook puts institutions in them too: four school
    # districts reached the roster this way and were hashed out of the data.
    # None of these tokens is plausible as a personal name in this corpus.
    if any(token.lower() in ORG_TOKENS for token in tokens):
        return False
    return True


def looks_like_given_name(part: str, organization_words: set[str]) -> bool:
    """Shape test for a bare first name, used ONLY in explicit person columns.

    `looks_like_person` requires two tokens, because in free text a lone
    capitalized word is far more often a sentence opener than a person. Inside a
    column headed `Contacts / Rounds` that reasoning inverts: a lone capitalized
    word there is a person by default, and requiring a surname leaked thirteen
    real interviewer names into a public repository.

    Organizations are excluded by WORD rather than by substring. Substring
    matching blocks "Heath" because the corpus holds "Solv Health", and a
    blocked name is a name shipped in the clear.
    """
    if " " in part or not part:
        return False
    if not part[:1].isupper() or part.isupper():
        return False
    # Letters, apostrophes and hyphens only. Stops "2025", "1x" and "per_ab12cd".
    if not re.fullmatch(r"[A-Z][A-Za-z'’\-]+", part):
        return False
    lowered = part.lower()
    if lowered in NON_NAME_WORDS or lowered in ORG_TOKENS:
        return False
    if part in KNOWN_NON_PERSONS:
        return False
    if lowered in organization_words:
        return False
    return True


def review_candidates(texts: list[str], organization_words: set[str]) -> set[str]:
    """Capitalized words that survived redaction and nothing has accounted for.

    The single-token prose case cannot be decided by shape: "Patrick originated
    the opportunity" and "Community post is the source" are the same shape. So it
    is decided by review, once, and this guard makes the review mandatory.

    It runs over the REDACTED text, not the source. What matters is what actually
    ships, and a name that the roster already replaced needs no review. Scanning
    the input instead flags every name the redaction handled correctly, which
    buries the real leaks in noise.
    """
    seen: set[str] = set()
    for text in texts:
        for word in re.findall(r"\b[A-Z][a-z]{2,}\b", text):
            if word in PROSE_GIVEN_NAMES or word in PROSE_NON_NAMES:
                continue
            if word.lower() in organization_words:
                continue
            if word.lower() in NON_NAME_WORDS or word.lower() in ORG_TOKENS:
                continue
            seen.add(word)
    return seen


def build_person_roster(wb) -> tuple[dict[str, str], set[str], list[str]]:
    """Names to redact, gathered from the workbook rather than hardcoded.

    Two harvests with different rules.

    EXPLICIT COLUMNS hold people by definition, so only shape is checked. A
    company that lands in one of them is over-redacted, which is the safe error.

    PROSE COLUMNS are free text where no split rule isolates a name, so a
    capitalized bigram is a candidate and organization tokens are filtered out.
    Missing a person here is no worse than before this harvest existed, while
    hashing a company name would destroy data the analysis needs.
    """
    organizations: set[str] = set()
    for sheet, column in (
        ("Combined Ledger", "Company"),
        ("Remaining Excluded", "Company"),
        ("LinkedIn Applications", "Company"),
        ("Jobright Applications", "Company"),
    ):
        if sheet not in wb.sheetnames:
            continue
        _, rows = sheet_rows(wb, sheet)
        for row in rows:
            value = clean(row.get(column))
            # `UNKNOWN — Jacob Bowman's company` is not a company name, it is a
            # description of a company nobody could identify, and several of them
            # are built out of the person's own name. Admitting those to the
            # organization set lets a name protect itself from redaction, which
            # is exactly how Jacob Bowman shipped in the clear.
            if value and not value.startswith("UNKNOWN"):
                organizations.add(value)

    # This repository's own company names count as organizations too. The
    # workbook's Company column only lists companies it KEPT, so every company it
    # dropped is invisible to the guard: Lumenalta, Proofpoint and Designit were
    # removed from its ledger, appear only in a prose note about their removal,
    # and were hashed as people. This census still holds all three, so it can say
    # what they are.
    for row in load_csv_column("adjudication/applications__full_census.csv", ("company_canonical", "company_as_listed")):
        organizations.add(row)

    # Word-level view of the same set, for the single-token test. See
    # looks_like_given_name for why substring matching is wrong there.
    organization_words = {w.lower().strip(".,()") for org in organizations for w in org.split()}

    roster: set[str] = set()
    # Given names of people confirmed by an explicit person column. Prose refers
    # to them by first name only, "the message to Andrew", "Jorge was a TA", and
    # the full-name roster never fires on those.
    given: set[str] = set()
    person_texts: list[str] = []

    for sheet, column in (
        ("LinkedIn Job Threads", "Person"),
        ("Uncertain and Quarantined", "Contact"),
        ("Interview & Opportunity Register", "Contacts / Rounds"),
    ):
        if sheet not in wb.sheetnames:
            continue
        _, rows = sheet_rows(wb, sheet)
        for row in rows:
            value = clean(row.get(column))
            person_texts.append(value)

            # Parenthetical content is removed from the WHOLE value before any
            # split. Splitting first breaks on separators that live inside the
            # parentheses: "Eddie (2 interviews; final was 1 hour)" split on `;`
            # yields "Eddie (2 interviews", an unbalanced fragment that no shape
            # test matches, and the name ships in the clear.
            # `+` is a separator too. Without it "Gurjap Sandhu + Kofi Boamah O."
            # stays one six-token fragment that no shape test matches, and two
            # full names ship in the clear.
            for raw in re.split(r"→|/|;|,|\+|\band\b", re.sub(r"\([^)]*\)", " ", value)):
                candidate = " ".join(raw.split())
                if looks_like_person(candidate, organizations):
                    roster.add(candidate)
                    given.add(candidate.split()[0])
                elif looks_like_given_name(candidate, organization_words):
                    # A bare first name. The column says it is a person, so the
                    # absence of a surname is not evidence that it is not one.
                    # It joins the given set, not the global one: a lone name is
                    # ambiguous outside a person column.
                    given.add(candidate)

            # The ORIGINAL forms too. "Jim (Boris) Ryss" reduces to "Jim Ryss",
            # and rostering only the reduction leaves the original written out in
            # full, which is how a real name shipped unredacted once already.
            for raw in re.split(r"→|/|;|,|\band\b", value):
                original = " ".join(raw.split())
                if "(" in original and ")" in original:
                    inner = " ".join(re.sub(r"\([^)]*\)", " ", original).split())
                    if inner and (
                        looks_like_person(inner, organizations)
                        or looks_like_given_name(inner, organization_words)
                    ):
                        roster.add(original)

    prose_name = re.compile(r"\b[A-Z][a-z]+(?:\s+[A-Z][a-z]+){1,2}\b")
    for sheet, column in (
        ("Interview & Opportunity Register", "Organization / Context"),
        ("Interview & Opportunity Register", "Origin"),
        ("Interview & Opportunity Register", "Reconciliation Notes"),
        ("Data Quality Findings", "Evidence"),
    ):
        if sheet not in wb.sheetnames:
            continue
        _, rows = sheet_rows(wb, sheet)
        for row in rows:
            value = clean(row.get(column))
            person_texts.append(value)
            for candidate in prose_name.findall(value):
                candidate = " ".join(candidate.split())
                if looks_like_person(candidate, organizations):
                    roster.add(candidate)
                    given.add(candidate.split()[0])

    # Reviewed single-token names from prose, where shape cannot decide.
    given |= PROSE_GIVEN_NAMES
    # A given name is only kept when it is not also a company word, so a person
    # called Every cannot take the company Every out of the data with them.
    given = {g for g in given if g.lower() not in organization_words}

    return (
        {name: person_pointer(name) for name in sorted(roster, key=len, reverse=True)},
        {name: person_pointer(name) for name in sorted(given, key=len, reverse=True)},
        organization_words,
        person_texts,
    )


def is_structural_column(header: str) -> bool:
    """True for columns holding a company, role, location or join key.

    Bare given names are NOT redacted in these. "Austin" is an interviewer at
    Every and also the city in "SDR Manager (Austin; relocation available)" and
    in two Jobright locations. Redacting the given name everywhere hashes the
    city out of a role title, which destroys data to protect nobody, since the
    person is already covered wherever they are actually named.

    Full names stay global. "Teresa Vitale" is distinctive enough to redact
    anywhere it appears; "Austin" is not.
    """
    lowered = header.strip().lower()
    if lowered.endswith("key"):
        return True
    return any(word in lowered for word in ("company", "role", "title", "location"))


def redact(text: str, roster: dict[str, str], given: dict[str, str] | None = None) -> str:
    """Replace every rostered personal name with its stable pointer.

    `given` holds bare first names and is passed only for columns where a lone
    capitalized word means a person. See is_structural_column.
    """
    for name, pointer in roster.items():
        if name and name in text:
            text = text.replace(name, pointer)
    for name, pointer in (given or {}).items():
        # Word-bounded: without it "Chris" rewrites the middle of "Christina"
        # and leaves a pointer glued to a name fragment.
        text = re.sub(rf"\b{re.escape(name)}\b", pointer, text)
    # Residual pattern: "Recruiter outbound: Firstname Lastname / Company".
    # Catches names that never appear in a dedicated person column.
    def _sub(match: re.Match) -> str:
        return f"{match.group(1)}{person_pointer(match.group(2))}{match.group(3)}"

    return re.sub(
        r"((?:outbound|outreach|intermediary|inbound|via|follow-up)[:\s]+)"
        r"([A-Z][a-z]+(?:\s+[A-Z][A-Za-z.'’-]+)+)"
        r"(\s*(?:/|→|$))",
        _sub,
        text,
    )


def write_csv(
    path: Path,
    header: list[str],
    rows: list[dict],
    roster: dict[str, str],
    given: dict[str, str],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = [h for h in header if h]
    scoped = {f: ({} if is_structural_column(f) else given) for f in fields}
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({f: redact(clean(row.get(f)), roster, scoped[f]) for f in fields})
    print(f"wrote {path.relative_to(ROOT)} ({len(rows)} rows)")


# sheet name -> output path. The LinkedIn export is the only one that lands in
# artifacts/, because it is the only one that is an artifact.
SHEETS = {
    "LinkedIn Applications": ARTIFACTS / "linkedin_job_applications_export.csv",
    "Combined Ledger": OUT / "checkpoint__ledger.csv",
    "Source Classification": OUT / "checkpoint__source_classification.csv",
    "Role Classification": OUT / "checkpoint__role_classification.csv",
    "Role Lane Distribution": OUT / "checkpoint__role_lane_distribution.csv",
    "LinkedIn Job Threads": OUT / "checkpoint__linkedin_threads.csv",
    "Jobright Applications": OUT / "checkpoint__jobright.csv",
    "Dedup Decisions": OUT / "checkpoint__dedup_decisions.csv",
    "Data Quality Findings": OUT / "checkpoint__data_quality_findings.csv",
    "Platform Merge Summary": OUT / "checkpoint__merge_summary.csv",
    "Uncertain and Quarantined": OUT / "checkpoint__uncertain.csv",
    "Remaining Excluded": OUT / "checkpoint__remaining_excluded.csv",
    # Added in the refined workbook. It is the challenger adopting a two-register
    # split, and it carries interview rounds and contacts that exist nowhere else.
    "Interview & Opportunity Register": OUT / "checkpoint__interview_register.csv",
}


def main() -> int:
    if len(sys.argv) > 1:
        path = Path(sys.argv[1])
    elif os.environ.get("CHECKPOINT_XLSX"):
        path = Path(os.environ["CHECKPOINT_XLSX"])
    else:
        print(
            f"No workbook path given. Pass it as an argument or set CHECKPOINT_XLSX.\n"
            f"Expected {SOURCE_NAME}, sha256 {SOURCE_SHA256}.\n"
            f"The extracted CSVs are committed, so nothing downstream needs this step re-run."
        )
        return 0

    if not path.is_file():
        print(f"Workbook not found at {path}. Extracted CSVs under challenge/ are already committed.")
        return 0

    digest = hashlib.sha256(path.read_bytes()).hexdigest()
    if digest in KNOWN_SOURCES:
        print(f"reading {KNOWN_SOURCES[digest]}")
    else:
        print(
            f"WARNING: workbook sha256 is {digest}, which is not a version this\n"
            f"extractor has seen. Re-read challenge/SUPPLEMENTARY-LEDGER.md before\n"
            f"trusting any comparison built on it."
        )

    wb = load_workbook(path)
    roster, given, organization_words, person_texts = build_person_roster(wb)
    print(
        f"redacting {len(roster)} personal names to per_ pointers, "
        f"plus {len(given)} bare given names in person-bearing columns only"
    )

    # Review gate. Runs on the redacted text and BEFORE anything is written, so a
    # failure leaves no half-redacted CSVs behind for the next run to be judged
    # against.
    unclassified = review_candidates(
        [redact(text, roster, given) for text in person_texts], organization_words
    )
    if unclassified:
        raise SystemExit(
            "Redaction review required. These capitalized words survive redaction in\n"
            "person-bearing columns and are on no list, so the extractor cannot tell\n"
            "whether they are people. Classify every one into PROSE_GIVEN_NAMES\n"
            "(redact it) or PROSE_NON_NAMES (leave it), then re-run.\n\n"
            "  " + "\n  ".join(sorted(unclassified)) + "\n\n"
            "Nothing was written. Failing here is the point: the previous version made\n"
            "this call silently and shipped real names into a public repository."
        )

    for sheet, destination in SHEETS.items():
        if sheet not in wb.sheetnames:
            print(f"skipping {sheet}, absent from this workbook version")
            continue
        header, rows = sheet_rows(wb, sheet)
        write_csv(destination, header, rows, roster, given)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
